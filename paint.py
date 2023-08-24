# -*- coding:utf-8 -*-
# author:skipper 
# date:8/12/23 
# time:8:27 PM 
# software: PyCharm

import openpyxl
from matplotlib import pyplot as plt
import pyecharts.charts
from pyecharts.charts import Scatter, Line, Bar
import pyecharts.options as opts


def replace_run_time_out(lis1:list,lis2:list)-> [list,list]:
    for i in range(len(lis1)):
        # print(type(lis1[i]))
        if 'run' in str(lis1[i]):
            lis1[i] = 50
        if 'run' in str(lis2[i]):
            lis2[i] = 50
        # m = max(lis1[i],lis2[i])
        m = lis1[i] + lis2[i]
        lis1[i] = lis1[i]/m
        lis2[i] = lis2[i]/m
    return lis1,lis2


def paint_bar(names,lis0,lis1,title):
    bar = (
        Bar()
        .add_xaxis(names)
        .add_yaxis('08-19', lis0)
        .add_yaxis('08-20', lis1)
        .set_series_opts(stack='stack')
        # .reversal_axis()
        .set_global_opts(title_opts=opts.TitleOpts(title=title))
    )
    # line = (
    #     Line()
    #     .add_xaxis(names)
    #     .add_yaxis('相等', [0.5 for i in range(len(names))])
    # )

    # bar.overlap(line).render("loop_parallel.html")
    bar.render("{}.html".format(title))


if __name__ == '__main__':
    file = './result.xlsx'
    wb = openpyxl.load_workbook(file)
    sheet = wb['Sheet2']
    size = sheet.max_row
    # print(size)
    names = []
    lis0 = []
    lis1 = []
    for i in range(2, size - 1):
        if 'False' in str(sheet.cell(i, 3).value):
            continue
        names.append(sheet.cell(i, 1).value)
        lis0.append(sheet.cell(i, 2).value)
        lis1.append(sheet.cell(i, 3).value)
    # print(lis0)
    lis0, lis1 = replace_run_time_out(lis0, lis1)
    x = [i for i in range(len(names))]
    title = "compare"
    paint_bar(names,lis0,lis1,title)
'''
"axisLabel": {
            // 设置标签旋转为45度
                rotate: -60,
                fontSize: 8
            },
'''